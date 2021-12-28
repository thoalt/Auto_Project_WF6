import softest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from utilites import global_dir
import os

from utilites.handle_config import ConfigParserIni

# col_TestID = 2
# col_TestInput = 5
# col_TestExpect = 6

configInfo = ConfigParserIni.config_data_info()

excel_file = configInfo['file_name']
col_TestID = int(configInfo['col_testid'])
col_TestInput = int(configInfo['col_testinput'])
col_TestExpect = int(configInfo['col_testexpect'])
row_start = 11
class ExcelPaser(softest.TestCase):
    @staticmethod
    def read_data_from_exel(excel_file, excel_sheet):
        dataList = []
        excel_path = os.path.join(global_dir.DATA_FILES_PATH, excel_file)
        wb = load_workbook(filename=excel_path)
        sh: Worksheet = wb[excel_sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(2, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            dataList.append(row)
        wb.close()
        return dataList

    @staticmethod
    def read_data_via_testcaseID(excel_sheet, testcaseID):
        data_lst = []
        data_input_lst = []
        dataExpect = ''
        dataInput = ''
        excel_path = os.path.join(global_dir.DATA_FILES_PATH, excel_file)
        wb = load_workbook(filename=excel_path)
        sh: Worksheet = wb[excel_sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        # Get data from excel file
        for rowID in range(row_start, row_ct):
            if testcaseID == sh.cell(rowID, col_TestID).value:
                dataInput = sh.cell(rowID, col_TestInput).value
                dataExpect = sh.cell(rowID, col_TestExpect).value
                break
        else:
            raise Exception("Cannot Find Testcase ID: %s in sheet %s" %(testcaseID, excel_sheet))

        if ('\n' in dataInput) and ('  ' not in dataInput):
            data_lst = dataInput.split("\n")
        elif ('\n' not in dataInput) and ('  ' in dataInput):
            data_lst = dataInput.split('  ')
        elif ('\n' in dataInput) and ('  ' in dataInput):
            rList = dataInput.split('\n')
            for lt in rList:
                data_lst.extend(lt.split('  '))
        else:
            data_lst = [dataInput]

        for ele in data_lst:
            #data_val = []
            if '(space)' in ele:
                strNew = ele.replace('(space)', ' ')
            elif ('trống') in ele:
                strNew = ele.replace('(trống)', '')
            else:
                strNew = ele
            data_input_lst.append([strNew, dataExpect])
        return data_input_lst

# paseExcel = ExcelPaser()
# data = paseExcel.read_data_via_testcaseID("RuleInputDefault", "RULE_FACTORY_LOCATION_03")
# print("Data get: " + data)